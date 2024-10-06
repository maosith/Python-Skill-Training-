import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

print("Cadastro de Crédito Pessoal")
opcao_1 = input("Deseja cadastrar um cliente novo? Se sim, digite 1. ")
if opcao_1 == '1':
    nome_aten = (input("Qual o nome do atendente? "))
    nome = input("Qual o nome do cliente? ")
    cpf = input("Qual o CPF do cliente? Sem pontos ou traços ")
    num_phone = input("Qual o número de contato do cliente? Sem traços. ")
    cred_p = input("Qual o valor oferecido pelo banco? Utilize ponto e virgula para digitar o valor, por exemplo: 1.000,00 ")
else:
    print("Saindo do sistema de cadastro.")


data_atual = datetime.now().strftime("%d/%m/%Y %H:%M")

df = pd.DataFrame({
    'Data': [data_atual],
    'Nome': [nome],
    'CPF': [cpf],
    'Tel': [num_phone],
    'Crédito': [cred_p],
    'Atendt': [nome_aten]
})

file_path = 'cadastro_clientes.xlsx'
try:
    book = load_workbook(file_path)

    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)

except FileNotFoundError:
    df.to_excel(file_path, index=False)

print(f"Cliente {nome} cadastrado com sucesso!")

recom_1 = int(input("Se deseja fazer mais um cadastro, digite 2 "))

if recom_1 == '2':
    print("Cadastro de Crédito Pessoal")
opcao_1 = input("Deseja cadastrar um cliente novamente? Se sim, digite 1. ")
if opcao_1 == '1':
    nome_aten = (input("Qual o nome do atendente? "))
    nome = input("Qual o nome do cliente? ")
    cpf = input("Qual o CPF do cliente? Sem pontos ou traços ")
    num_phone = input("Qual o número de contato do cliente? Sem traços. ")
    cred_p = input("Qual o valor oferecido pelo banco? Utilize ponto e virgula para digitar o valor, por exemplo: 1.000,00 ")
else:
    print("Saindo do sistema de cadastro.")


data_atual = datetime.now().strftime("%d/%m/%Y %H:%M")

df = pd.DataFrame({
    'Data': [data_atual],
    'Nome': [nome],
    'CPF': [cpf],
    'Tel': [num_phone],
    'Crédito': [cred_p],
    'Atendt': [nome_aten]
})

file_path = 'cadastro_clientes.xlsx'
try:
    book = load_workbook(file_path)

    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)

except FileNotFoundError:
    df.to_excel(file_path, index=False)

print(f"Cliente {nome} cadastrado com sucesso!")